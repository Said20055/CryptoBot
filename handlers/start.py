"""
Обработчики для основных команд бота.

Этот модуль содержит обработчики для команд /start, /help, /id,
а также функции для работы с пользователями и рассылками.
"""

import os
from aiogram.types import (
    Message,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    BufferedInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from config import ADMIN_CHAT_ID
from utils.logging_config import logger
from utils.texts import WELCOME_PHOTO_URL, WELCOME_TEXT, MAIN_KEYBOARD, format_user_display_name
from utils.database import save_or_update_user, find_all_users
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Определяем состояния для рассылки
class BroadcastStates(StatesGroup):
    waiting_for_message = State()

async def start_handler(event: Message | CallbackQuery):
    """Обработчик команды /start и возврата в главное меню.
    
    Args:
        event: Сообщение или callback query
    """
    photo_url = WELCOME_PHOTO_URL
    text = WELCOME_TEXT
    keyboard = MAIN_KEYBOARD
    if isinstance(event, Message):
        username = event.from_user.username or "Нет username"
        full_name = event.from_user.full_name
        user_id = event.from_user.id
        is_new_user = await save_or_update_user(user_id, username, full_name)
        logger.info(f"User {user_id} ({username}) {'saved as new' if is_new_user else 'updated'} in DB")
        # Отправляем уведомление админам при каждом /start
        for admin_id in ADMIN_CHAT_ID:
            try:
                await event.bot.send_message(
                    chat_id=admin_id,
                    text=f"Пользователь запустил /start:\nUsername: {format_user_display_name(username)}\nПолное имя: {full_name}\nID: {user_id}\nНовый: {is_new_user}",
                    parse_mode=None
                )
                logger.info(f"Notification sent to admin {admin_id} about user {username}")
            except (ConnectionError, TimeoutError) as e:
                logger.error(f"Failed to notify admin {admin_id} about user {user_id}: {e}")
        await event.answer_photo(
            photo=photo_url, caption=text, reply_markup=keyboard, parse_mode="Markdown"
        )
    else:  # isinstance(event, CallbackQuery)
        logger.info(f"User {event.from_user.id} returned to main menu via callback")
        await event.message.answer_photo(
            photo=photo_url, caption=text, reply_markup=keyboard, parse_mode="Markdown"
        )
        await event.message.delete()

async def sell_handler(callback_query: CallbackQuery):
    """Обработчик кнопки продажи криптовалюты.
    
    Args:
        callback_query: Callback query от нажатия кнопки
    """
    await callback_query.message.answer(
        "Продажа криптовалюты. Напишите оператору ваши детали сделки или ждите дальнейших шагов.")
    await callback_query.answer()

async def buy_handler(callback_query: CallbackQuery):
    """Обработчик кнопки покупки криптовалюты.
    
    Args:
        callback_query: Callback query от нажатия кнопки
    """
    await callback_query.message.answer(
        "Покупка криптовалюты. Напишите оператору ваши детали сделки или ждите дальнейших шагов.")
    await callback_query.answer()

async def operator_handler(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик кнопки связи с оператором.
    
    Args:
        callback_query: Callback query от нажатия кнопки
        state: FSM состояние
    """
    from utils.states import TransactionStates
    from config import ADMIN_CHAT_ID
    
    # Получаем данные из состояния
    data = await state.get_data()
    action = data.get('action')
    crypto = data.get('crypto')
    amount_crypto = data.get('amount_crypto')
    amount_rub = data.get('amount_rub')
    total_amount = data.get('total_amount', amount_rub)
    
    if action and crypto and amount_crypto and amount_rub:
        # Если есть данные о транзакции, создаем заявку для оператора
        username = callback_query.from_user.username or callback_query.from_user.full_name or "Пользователь"
        user_id = callback_query.from_user.id
        
        # Определяем что запрашивать в зависимости от типа операции
        if action == 'sell':
            request_text = "Укажите ваш номер телефона и название банка для получения рублей\nПример: +79999999999, Сбербанк"
        else:
            request_text = f"Укажите адрес вашего {crypto} кошелька для получения криптовалюты\nПример: bc1qtmgxuhpqt5l36pwcrnh9ujkm0afgrqvgumnk0n"
        
        text = (
            f"👨‍💼 Связь с оператором\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"💳 {action.title()} {crypto}\n"
            f"💎 Сумма: {amount_crypto:,.8f} {crypto}\n"
            f"💰 Сумма в рублях: {total_amount:,.2f} ₽\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{request_text}"
        )
        
        # Уведомляем админов о запросе оператора
        from utils.texts import format_user_display_name
        admin_text = (
            f"👤 *Пользователь запросил оператора:* {format_user_display_name(username)} (ID: {user_id})\n"
            f"💳 Операция: {action.title()} {crypto}\n"
            f"💎 Сумма: {amount_crypto:,.8f} {crypto}\n"
            f"💰 Сумма в рублях: {total_amount:,.2f} ₽\n\n"
            f"Свяжитесь с пользователем для завершения сделки."
        )
        
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        admin_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="💬 Ответить пользователю", callback_data=f"admin_reply_{user_id}")]
            ]
        )
        
        for admin_id in ADMIN_CHAT_ID:
            try:
                await callback_query.bot.send_message(
                    chat_id=admin_id,
                    text=admin_text,
                    reply_markup=admin_keyboard,
                    parse_mode="Markdown"
                )
                logger.info(f"Operator request sent to admin {admin_id}")
            except (ConnectionError, TimeoutError) as e:
                logger.error(f"Failed to send operator request to admin {admin_id}: {e}")
        
        await callback_query.message.reply(text, parse_mode="Markdown")
        await state.set_state(TransactionStates.waiting_for_phone)
    else:
        # Если нет данных о транзакции, просто показываем контакт оператора
        await callback_query.message.answer(
            "👨‍💼 Связь с оператором: @jenya2hh\n\n"
            "Напишите оператору для получения помощи с обменом криптовалют.")
    
    await callback_query.answer()

async def reviews_handler(callback_query: CallbackQuery):
    """Обработчик кнопки отзывов.
    
    Args:
        callback_query: Callback query от нажатия кнопки
    """
    await callback_query.message.answer(
        "Отзывы о нашем сервисе: скоро добавим ссылку или отправим подборку.")
    await callback_query.answer()

async def help_handler(message: Message):
    """Обработчик команды /help.
    
    Args:
        message: Сообщение с командой /help
    """
    text = (
        "📚 *Справка по боту*\n\n"
        "Я — бот Обмен криптовалют Вот что я умею:\n"
        "- /start — Запустить бота и открыть главное меню.\n"
        "- /id — Получить ваш Telegram ID.\n"
        "Выберите опцию в меню или используйте команды!"
    )
    await message.answer(text, parse_mode="Markdown")

async def id_handler(message: Message):
    """Обработчик команды /id.
    
    Args:
        message: Сообщение с командой /id
    """
    user_id = message.from_user.id
    await message.answer(f"Ваш Telegram ID: {user_id}")
    logger.info(f"User {user_id} requested their Telegram ID")

async def image_handler(message: Message):
    """Обработчик изображений.
    
    Args:
        message: Сообщение с изображением
    """
    if message.photo:
        photo_id = message.photo[-1].file_id
        username = message.from_user.username or message.from_user.full_name or "Пользователь"
        user_id = message.from_user.id
        
        logger.info(f"Received photo with ID: {photo_id} from user {user_id}")
        
        # Подтверждаем пользователю
        await message.answer("✅ Изображение получено и отправлено оператору.")
        
        # Отправляем изображение админам
        admin_text = (
            f"📷 *Изображение от пользователя:* {format_user_display_name(username)} (ID: {user_id})\n\n"
            f"Пользователь отправил изображение. Ответьте ему, используя кнопку ниже."
        )
        
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        admin_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="💬 Ответить пользователю", callback_data=f"admin_reply_{user_id}")]
            ]
        )
        
        for admin_id in ADMIN_CHAT_ID:
            try:
                await message.bot.send_photo(
                    chat_id=admin_id,
                    photo=photo_id,
                    caption=admin_text,
                    reply_markup=admin_keyboard,
                    parse_mode="Markdown"
                )
                logger.info(f"Photo sent to admin {admin_id}")
            except (ConnectionError, TimeoutError) as e:
                logger.error(f"Failed to send photo to admin {admin_id}: {e}")

async def users_handler(message: Message):
    """Обработчик команды /users для получения списка пользователей.
    
    Args:
        message: Сообщение с командой /users
    """
    user_id = message.from_user.id
    if user_id not in ADMIN_CHAT_ID:
        await message.answer("Эта команда доступна только администраторам.")
        logger.info(f"User {user_id} attempted to use /users but is not an admin")
        return

    try:
        users = await find_all_users()
        if not users:
            await message.answer("Пользователи не найдены в базе данных.")
            logger.info("No users found in database for /users command")
            return

        # Создаём DataFrame из данных пользователей, исключая invite_link_issued и subscription_duration
        columns = ["user_id", "username", "full_name", "email", "is_subscribed", "subscription_end", "created_at"]
        df = pd.DataFrame(users, columns=["user_id", "username", "full_name", "email", "is_subscribed", "subscription_end", "created_at", "invite_link_issued", "subscription_duration"])
        df = df[columns]  # Убираем ненужные столбцы
        
        # Преобразуем столбцы для удобства чтения
        df["is_subscribed"] = df["is_subscribed"].apply(lambda x: "Да" if x else "Нет")
        # Преобразуем даты в строки, обрабатывая NaT и None
        df["subscription_end"] = pd.to_datetime(df["subscription_end"], errors='coerce').dt.strftime("%d.%m.%Y %H:%M:%S").fillna("Нет")
        df["created_at"] = pd.to_datetime(df["created_at"], errors='coerce').dt.strftime("%d.%m.%Y %H:%M:%S").fillna("Нет")

        # Создаём Excel-файл в памяти
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Users")
            worksheet = writer.sheets["Users"]
            
            # Определяем цвета и стили
            header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")  # Синий (SteelBlue)
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый
            gray_fill = PatternFill(start_color="EAEAE9", end_color="EAEAE9", fill_type="solid")  # Светло-серый (#EAEAE9)
            header_font = Font(color="FFFFFF", bold=True)  # Белый текст, жирный шрифт
            header_alignment = Alignment(horizontal="center", vertical="center")
            # Определяем границы
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            
            # Применяем стили к заголовкам
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # Настраиваем ширину столбца
                max_length = max(
                    df[column].astype(str).map(len).max(),  # Максимальная длина содержимого
                    len(str(column))  # Длина заголовка
                )
                adjusted_width = max_length + 2
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
            
            # Чередование цветов строк и добавление границ
            for row_num in range(2, len(df) + 2):
                fill = gray_fill if (row_num % 2 == 0) else white_fill
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    cell.border = thin_border
            
            # Добавляем фильтры для сортировки
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        output.seek(0)
        
        # Отправляем файл админу
        file = BufferedInputFile(output.getvalue(), filename="users.xlsx")
        await message.answer_document(file, caption="Список пользователей из базы данных")
        logger.info(f"User {user_id} received users list as Excel file")
        
    except (IOError, ValueError, KeyError) as e:
        await message.answer("Ошибка при генерации списка пользователей. Пожалуйста, попробуйте позже.")
        logger.error(f"Failed to generate users list for user {user_id}: {e}")

async def log_handler(message: Message):
    """Обработчик команды /log для получения файла логов.
    
    Args:
        message: Сообщение с командой /log
    """
    user_id = message.from_user.id
    if user_id not in ADMIN_CHAT_ID:
        await message.answer("Эта команда доступна только администраторам.")
        logger.info(f"User {user_id} attempted to use /log but is not an admin")
        return

    log_file_path = "logging.log"
    try:
        if not os.path.exists(log_file_path):
            await message.answer("Файл логов не найден.")
            logger.info("Log file not found for /log command")
            return

        with open(log_file_path, "rb") as f:
            file = BufferedInputFile(f.read(), filename="bot.log")
            await message.answer_document(file, caption="Файл логов бота")
            logger.info(f"User {user_id} received log file")
    except (IOError, FileNotFoundError) as e:
        await message.answer("Ошибка при отправке файла логов. Пожалуйста, попробуйте позже.")
        logger.error(f"Failed to send log file to user {user_id}: {e}")

async def send_handler(message: Message, state: FSMContext):
    """Обработчик команды /send для рассылки сообщений.
    
    Args:
        message: Сообщение с командой /send
        state: FSM состояние
    """
    user_id = message.from_user.id
    if user_id not in ADMIN_CHAT_ID:
        await message.answer("Эта команда доступна только администраторам.")
        logger.info(f"User {user_id} attempted to use /send but is not an admin")
        return

    await message.answer(
        "Пожалуйста, отправьте текст для рассылки.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Отменить", callback_data="cancel_broadcast")]
            ]
        )
    )
    await state.set_state(BroadcastStates.waiting_for_message)

async def cancel_broadcast_handler(callback_query: CallbackQuery, state: FSMContext):
    """Обработчик отмены рассылки.
    
    Args:
        callback_query: Callback query от нажатия кнопки "Отменить"
        state: FSM состояние
    """
    user_id = callback_query.from_user.id
    if user_id not in ADMIN_CHAT_ID:
        await callback_query.message.answer("Эта команда доступна только администраторам.")
        logger.info(f"User {user_id} attempted to cancel broadcast but is not an admin")
        await callback_query.message.delete()
        return

    await state.clear()
    await callback_query.message.answer("Рассылка отменена.")
    logger.info(f"User {user_id} cancelled broadcast")
    await callback_query.message.delete()

async def broadcast_message_handler(message: Message, state: FSMContext):
    """Обработчик сообщения для рассылки.
    
    Args:
        message: Сообщение с текстом для рассылки
        state: FSM состояние
    """
    user_id = message.from_user.id
    if user_id not in ADMIN_CHAT_ID:
        await message.answer("Эта команда доступна только администраторам.")
        logger.info(f"User {user_id} attempted to broadcast message but is not an admin")
        await state.clear()
        return

    broadcast_text = message.text
    try:
        users = await find_all_users()
        if not users:
            await message.answer("Пользователи не найдены в базе данных.")
            logger.info("No users found for broadcast")
            await state.clear()
            return

        successful = 0
        failed = 0
        failed_users = []

        for user in users:
            target_user_id = user[0]
            username = user[1] or user[2] or "Пользователь"  # user[1] = username, user[2] = full_name
            try:
                await message.bot.send_message(
                    chat_id=target_user_id,
                    text=broadcast_text,
                    parse_mode="Markdown"
                )
                successful += 1
                logger.info(f"Broadcast message sent to user {target_user_id} ({format_user_display_name(username)})")
            except (ConnectionError, TimeoutError) as e:
                failed += 1
                failed_users.append(f"{format_user_display_name(username)} (ID: {target_user_id})")
                logger.error(f"Failed to send broadcast to user {target_user_id}: {e}")

        # Отправляем отчёт админу, вызвавшему команду
        report = (
            f"Рассылка завершена.\n"
            f"Отправлено: {successful}\n"
            f"Не отправлено: {failed}\n"
        )
        if failed_users:
            report += "Не удалось отправить следующим пользователям:\n" + "\n".join(failed_users)
        await message.answer(report)
        logger.info(f"Broadcast report sent to admin {user_id}: {successful} successful, {failed} failed")

        # Уведомляем всех админов об отчёте
        for admin_id in ADMIN_CHAT_ID:
            if admin_id != user_id:  # Не отправляем отчёт админу, который инициировал рассылку
                try:
                    await message.bot.send_message(
                        chat_id=admin_id,
                        text=report,
                        parse_mode=None
                    )
                    logger.info(f"Broadcast report sent to admin {admin_id}")
                except (ConnectionError, TimeoutError) as e:
                    logger.error(f"Failed to send broadcast report to admin {admin_id}: {e}")

        await state.clear()
    except (ConnectionError, TimeoutError, ValueError) as e:
        await message.answer("Ошибка при выполнении рассылки. Пожалуйста, попробуйте позже.")
        logger.error(f"Failed to perform broadcast for user {user_id}: {e}")
        await state.clear()